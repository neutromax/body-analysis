"""
Auto Annotator — Excel Writer
===============================

Appends aggregated annotation rows to an existing .xlsx file.

Output format (Section 7):
    Video Link | Time Window | Hand | Leg | Head | Body

Rules:
    - The video link is written ONCE on the first row of that video's
      block; subsequent rows for the same video leave that cell blank.
    - Never overwrites existing rows — always appends below.
    - Uses openpyxl for append operations (not pandas to_excel, which
      would overwrite the file).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from auto_config import EXCEL_COLUMNS, ENCODING

logger = logging.getLogger(__name__)


class ExcelWriter:
    """
    Manages appending rows to a user-selected .xlsx file.

    Usage:
        writer = ExcelWriter("output.xlsx")
        writer.start_video("https://youtube.com/watch?v=...")
        writer.append_row("0:00 – 0:10", {"hand": 1, "leg": 0, "head": -1, "body": 0})
        writer.append_row("0:10 – 0:20", {"hand": 0, "leg": 0, "head": 0, "body": 1})
    """

    def __init__(self, filepath: str) -> None:
        """
        Open or create the Excel workbook.

        If the file exists, opens it for appending.
        If it doesn't exist, creates it with the correct header row.

        Args:
            filepath:  Path to the .xlsx file.
        """
        self.filepath = Path(filepath)
        self._video_url: Optional[str] = None
        self._is_first_row_of_video: bool = True
        self._rows_written: int = 0

        if self.filepath.exists():
            self._wb = openpyxl.load_workbook(str(self.filepath))
            # Use the first sheet, or create one if empty.
            if self._wb.sheetnames:
                self._ws = self._wb.active
            else:
                self._ws = self._wb.create_sheet("Data")
                self._write_header()
            logger.info("Opened existing workbook: %s", self.filepath)
        else:
            self._wb = openpyxl.Workbook()
            self._ws = self._wb.active
            self._ws.title = "Data"
            self._write_header()
            self._save()
            logger.info("Created new workbook: %s", self.filepath)

    def _write_header(self) -> None:
        """Write the column header row if the sheet is empty."""
        if self._ws.max_row is None or self._ws.max_row < 1:
            for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
                self._ws.cell(row=1, column=col_idx, value=col_name)
        else:
            # Check if header already exists.
            first_cell = self._ws.cell(row=1, column=1).value
            if first_cell != EXCEL_COLUMNS[0]:
                for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
                    self._ws.cell(row=1, column=col_idx, value=col_name)

    def _save(self) -> None:
        """Save the workbook to disk."""
        try:
            self._wb.save(str(self.filepath))
            if getattr(self, "_save_failed_previously", False):
                logger.info("Excel file successfully saved after previous lock was released.")
                self._save_failed_previously = False
        except PermissionError:
            self._save_failed_previously = True
            logger.warning(
                "Excel file is locked (possibly open in another program like Excel). "
                "The row was saved in-memory and will be written to disk on the next update once the file is closed: %s",
                self.filepath,
            )
        except OSError as e:
            logger.error("Failed to save Excel file: %s", e)
            raise

    def start_video(self, video_url: str) -> None:
        """
        Signal that a new video is starting.

        The URL will be written in the first row of this video's block.
        Subsequent rows will leave the Video Link cell blank.

        Args:
            video_url:  The YouTube URL being annotated.
        """
        self._video_url = video_url
        self._is_first_row_of_video = True
        logger.info("Starting video block: %s", video_url)

    def append_row(
        self,
        time_window: str,
        result: dict[str, Optional[int]],
    ) -> int:
        """
        Append one aggregated row to the Excel file.

        Args:
            time_window:  Human-readable time range (e.g. "0:00 – 0:10").
            result:       Dict with keys "hand", "leg", "head", "body",
                          each mapped to -1, 0, 1, or None.

        Returns:
            Total number of rows written in this session.
        """
        next_row = self._ws.max_row + 1

        # Column 1: Video Link (only on the first row of this video).
        if self._is_first_row_of_video and self._video_url:
            self._ws.cell(row=next_row, column=1, value=self._video_url)
            self._is_first_row_of_video = False
        # else: leave the cell blank.

        # Column 2: Time Window
        self._ws.cell(row=next_row, column=2, value=time_window)

        # Columns 3–6: Hand, Leg, Head, Body
        for col_idx, feature in enumerate(["hand", "leg", "head", "body"], start=3):
            value = result.get(feature)
            if value is not None:
                self._ws.cell(row=next_row, column=col_idx, value=value)
            else:
                self._ws.cell(row=next_row, column=col_idx, value="NA")

        self._save()
        self._rows_written += 1

        logger.info(
            "Row %d written: %s → hand=%s, leg=%s, head=%s, body=%s",
            self._rows_written, time_window,
            result.get("hand"), result.get("leg"),
            result.get("head"), result.get("body"),
        )

        return self._rows_written

    @property
    def rows_written(self) -> int:
        """Total rows appended in this session."""
        return self._rows_written

    def close(self) -> None:
        """Close the workbook (final save)."""
        try:
            self._save()
        except Exception:
            pass  # Best effort on close.
        self._wb.close()
        logger.info("ExcelWriter closed. Total rows written: %d", self._rows_written)
