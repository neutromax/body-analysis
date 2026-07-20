"""
Auto Annotator — Video Queue Manager
======================================

Manages reading, writing, and tracking progress in the Link Queue Excel sheet.
Handles YouTube video ID extraction, status updates, and checkpoint verification.
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import openpyxl
import json

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = [
    "Video ID",
    "Video Link",
    "Status",
    "Last Second Completed",
    "Last Saved At",
]

def extract_video_id(url: str) -> str:
    """Extract 11-char video ID from various YouTube URL formats."""
    if not url:
        raise ValueError("URL is empty")
    patterns = [
        r"(?:v=|\/v\/|youtu\.be\/|\/embed\/)([a-zA-Z0-9_-]{11})",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    raise ValueError(f"Could not extract video ID from URL: {url}")


class VideoQueueManager:
    """
    Handles reading and writing the Link Queue Excel sheet.
    
    Exposes an interface to retrieve the next video, track progress,
    and cross-check progress against local JSON checkpoint files.
    """

    def __init__(self, queue_xlsx_path: str) -> None:
        self.filepath = Path(queue_xlsx_path)
        self._wb: Optional[openpyxl.Workbook] = None
        self._ws: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        self._col_map: Dict[str, int] = {}  # maps header name -> 1-based column index
        
        self.load_queue()

    def load_queue(self) -> None:
        """Open and validate the Link Queue sheet, mapping the header columns."""
        if not self.filepath.exists():
            raise FileNotFoundError(f"Link Queue Excel file not found: {self.filepath}")

        try:
            self._wb = openpyxl.load_workbook(str(self.filepath))
            # Use active sheet
            self._ws = self._wb.active
        except Exception as e:
            raise IOError(f"Could not open Link Queue Excel file: {e}")

        # Validate header row
        header_row = [cell.value for cell in self._ws[1]]
        missing = []
        for col_name in REQUIRED_COLUMNS:
            found = False
            for idx, cell_val in enumerate(header_row, start=1):
                if cell_val == col_name:
                    self._col_map[col_name] = idx
                    found = True
                    break
            if not found:
                missing.append(col_name)

        if missing:
            self._wb.close()
            raise ValueError(f"Link Queue Excel file is missing required column(s): {', '.join(missing)}")

    def _save(self) -> None:
        """Save Link Queue sheet back to disk."""
        if self._wb:
            try:
                self._wb.save(str(self.filepath))
                if getattr(self, "_save_failed_previously", False):
                    logger.info("Link Queue Excel file successfully saved after previous lock was released.")
                    self._save_failed_previously = False
            except PermissionError:
                self._save_failed_previously = True
                logger.warning(
                    "Link Queue Excel file is locked (possibly open in another program like Excel). "
                    "The progress was saved in-memory and will be written to disk on the next update once the file is closed: %s",
                    self.filepath,
                )

    def get_all_rows(self) -> List[Dict[str, Any]]:
        """Return a list of dicts representing all data rows in the queue."""
        rows = []
        max_row = self._ws.max_row
        if max_row < 2:
            return rows

        for r_idx in range(2, max_row + 1):
            row_data = {}
            for col_name, col_idx in self._col_map.items():
                row_data[col_name] = self._ws.cell(row=r_idx, column=col_idx).value
            row_data["row_idx"] = r_idx
            rows.append(row_data)
        return rows

    def get_next_video(self) -> Tuple[Optional[Dict[str, Any]], List[str]]:
        """
        Finds the first row whose Status is not 'Done'.
        Checks for duplicates and extracts missing video IDs if needed.
        
        Returns:
            A tuple of:
            - dict containing video details (or None if queue is complete)
            - list of string warning messages (e.g. duplicate warnings)
        """
        warnings = []
        all_rows = self.get_all_rows()
        processed_video_ids = set()
        duplicate_ids = set()

        # Step 1: Pre-process video IDs, fill in missing ones, check for duplicates
        changes_made = False
        for row in all_rows:
            r_idx = row["row_idx"]
            link = row["Video Link"]
            video_id = row["Video ID"]

            if not link:
                continue

            # Derive Video ID if blank
            if not video_id:
                try:
                    video_id = extract_video_id(link)
                    self._ws.cell(row=r_idx, column=self._col_map["Video ID"], value=video_id)
                    row["Video ID"] = video_id
                    changes_made = True
                    logger.info("Auto-derived Video ID '%s' for row %d", video_id, r_idx)
                except Exception as e:
                    # Mark malformed link as Error
                    self._set_error(r_idx, f"Malformed Link: {e}")
                    row["Status"] = "Error"
                    changes_made = True
                    continue

            # Check duplicates
            if video_id in processed_video_ids:
                duplicate_ids.add(video_id)
            else:
                processed_video_ids.add(video_id)

        # Save any derived IDs only if changes were made
        if changes_made:
            self._save()

        # Generate warnings for duplicates
        if duplicate_ids:
            warnings.append(
                f"Data Warning: Duplicate Video IDs detected in Link Queue: {', '.join(duplicate_ids)}. "
                f"Duplicate occurrences will be skipped."
            )
            for vid in duplicate_ids:
                logger.warning("Duplicate Video ID detected: %s. Skip rules will apply.", vid)

        # Step 2: Find the first non-Done row, skipping duplicates
        seen_ids = set()
        for row in all_rows:
            video_id = row["Video ID"]
            status = row["Status"]
            r_idx = row["row_idx"]

            if not video_id or status == "Error":
                continue

            # If it's a duplicate of a video we already processed (or will process), skip it
            if video_id in seen_ids:
                continue
            seen_ids.add(video_id)

            status_clean = str(status).strip() if status else "Not Started"
            if status_clean != "Done":
                # Check if this video has a duplicate row marked Done earlier, or if it is marked Done in Excel.
                # Actually, the spec says "A video whose Status is Done must never be reloaded or reprocessed,
                # even if the user pastes its link manually or it reappears elsewhere in the sheet."
                # So we check if ANY row with this video_id has status "Done".
                has_done_row = False
                for r in all_rows:
                    if r["Video ID"] == video_id and str(r["Status"]).strip() == "Done":
                        has_done_row = True
                        break
                
                if has_done_row:
                    logger.info("Video %s is marked Done in another row. Skipping.", video_id)
                    continue

                # Return the details of the active row
                return {
                    "row_idx": r_idx,
                    "video_id": video_id,
                    "video_link": row["Video Link"],
                    "status": status_clean,
                    "last_sec_completed": row["Last Second Completed"],
                }, warnings

        return None, warnings

    def _set_error(self, row_idx: int, reason: str) -> None:
        """Helper to mark a row's status as Error with a reason."""
        self._ws.cell(
            row=row_idx,
            column=self._col_map["Status"],
            value=f"Error: {reason}",
        )
        self._ws.cell(
            row=row_idx,
            column=self._col_map["Last Saved At"],
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    def cross_check_checkpoint(
        self,
        video_id: str,
        last_sec_completed_excel: Optional[int],
    ) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        """
        Verifies that the Excel sheet's Last Second Completed value agrees
        with what the local per-video JSON checkpoint file contains.

        Returns:
            (is_ok, message, checkpoint_data)
        """
        # Determine paths
        session_dir = Path(__file__).parent.parent / "data" / "sessions"
        checkpoint_path = session_dir / f"{video_id}.json"

        # Safe parsing of excel value
        try:
            excel_val = int(last_sec_completed_excel) if last_sec_completed_excel is not None else -1
        except (ValueError, TypeError):
            excel_val = -1

        if excel_val < 0:
            # Excel says Not Started (or blank)
            if checkpoint_path.exists():
                # Checkpoint exists! Read it to see if it contains observations
                try:
                    with open(checkpoint_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    obs_count = 0
                    max_time = -1
                    for win in data.get("windows", []):
                        for obs in win.get("observations", []):
                            if any(obs.get(k) is not None for k in ["hand", "leg", "head", "body"]):
                                obs_count += 1
                                max_time = max(max_time, obs.get("time_sec", -1))
                    
                    if obs_count > 0:
                        msg = (
                            f"Warning: Disagreement for video {video_id}. Excel says 'Not Started', "
                            f"but local checkpoint file '{checkpoint_path.name}' has {obs_count} "
                            f"annotated seconds (last second = {max_time})."
                        )
                        return False, msg, data
                except Exception as e:
                    pass
            return True, "", None

        # Excel says In Progress
        if not checkpoint_path.exists():
            msg = (
                f"Warning: Disagreement for video {video_id}. Excel says 'In Progress' (last completed "
                f"second: {excel_val}), but local checkpoint file '{checkpoint_path.name}' is missing."
            )
            return False, msg, None

        # Checkpoint exists, parse it and check max annotated second
        try:
            with open(checkpoint_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            max_annotated_sec = -1
            for win in data.get("windows", []):
                for obs in win.get("observations", []):
                    if any(obs.get(k) is not None for k in ["hand", "leg", "head", "body"]):
                        max_annotated_sec = max(max_annotated_sec, obs.get("time_sec", -1))

            if max_annotated_sec != excel_val:
                msg = (
                    f"Warning: Disagreement for video {video_id}. Excel completed second: {excel_val}, "
                    f"but local checkpoint file completed second: {max_annotated_sec}."
                )
                return False, msg, data
            
            return True, "", data
        except Exception as e:
            msg = f"Warning: Could not read or parse checkpoint file '{checkpoint_path.name}': {e}"
            return False, msg, None

    def mark_progress(self, video_id: str, second: int) -> None:
        """Finds row for video_id, updates status and last completed second, and saves."""
        # Find row by Video ID
        row_idx = self._find_row_by_video_id(video_id)
        if not row_idx:
            logger.warning("Could not mark progress: Video ID %s not found in queue", video_id)
            return

        self._ws.cell(row=row_idx, column=self._col_map["Status"], value="In Progress")
        self._ws.cell(row=row_idx, column=self._col_map["Last Second Completed"], value=second)
        self._ws.cell(
            row=row_idx,
            column=self._col_map["Last Saved At"],
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        self._save()
        logger.info("Marked progress in Excel queue for %s: second %d", video_id, second)

    def mark_done(self, video_id: str) -> None:
        """Finds row for video_id, sets status to Done, and saves."""
        row_idx = self._find_row_by_video_id(video_id)
        if not row_idx:
            logger.warning("Could not mark Done: Video ID %s not found in queue", video_id)
            return

        self._ws.cell(row=row_idx, column=self._col_map["Status"], value="Done")
        self._ws.cell(
            row=row_idx,
            column=self._col_map["Last Saved At"],
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        self._save()
        logger.info("Marked video %s as Done in Excel queue", video_id)

    def mark_error(self, video_id: str, reason: str) -> None:
        """Finds row for video_id, sets status to Error with reason, and saves."""
        row_idx = self._find_row_by_video_id(video_id)
        if not row_idx:
            logger.warning("Could not mark Error: Video ID %s not found in queue", video_id)
            return
        self._set_error(row_idx, reason)
        self._save()

    def _find_row_by_video_id(self, video_id: str) -> Optional[int]:
        """Finds the first row index matching the given Video ID (checking only non-duplicate/first occurrence)."""
        max_row = self._ws.max_row
        if max_row < 2:
            return None
        col_idx = self._col_map["Video ID"]
        for r in range(2, max_row + 1):
            val = self._ws.cell(row=r, column=col_idx).value
            if val and str(val).strip() == video_id:
                return r
        return None

    def is_video_done(self, video_id: str) -> bool:
        """Check if any row in the queue sheet for this video_id is marked Done."""
        for row in self.get_all_rows():
            if row.get("Video ID") == video_id and str(row.get("Status")).strip() == "Done":
                return True
        return False
