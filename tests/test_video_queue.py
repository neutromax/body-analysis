import os
import sys
import json
import pytest
import openpyxl
from pathlib import Path
from tempfile import TemporaryDirectory

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "auto_annotator"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from video_queue import VideoQueueManager, extract_video_id, REQUIRED_COLUMNS
from feature_classifier import Thresholds


def create_temp_xlsx(path: str, headers: list, data: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r_idx, row_vals in enumerate(data, start=2):
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=col_idx, value=val)
    wb.save(path)
    wb.close()


def test_extract_video_id():
    assert extract_video_id("https://www.youtube.com/watch?v=dQw4w9WgXcQ") == "dQw4w9WgXcQ"
    assert extract_video_id("https://youtu.be/dQw4w9WgXcQ") == "dQw4w9WgXcQ"
    assert extract_video_id("https://www.youtube.com/embed/dQw4w9WgXcQ") == "dQw4w9WgXcQ"
    with pytest.raises(ValueError):
        extract_video_id("https://example.com/not-youtube")


def test_load_queue_missing_columns():
    with TemporaryDirectory() as tmpdir:
        temp_excel = os.path.join(tmpdir, "bad_queue.xlsx")
        create_temp_xlsx(temp_excel, ["Video Link", "Status"], [["https://youtu.be/abc", "Not Started"]])
        with pytest.raises(ValueError) as exc:
            VideoQueueManager(temp_excel)
        assert "missing required column(s)" in str(exc.value)


def test_get_next_video_success():
    with TemporaryDirectory() as tmpdir:
        temp_excel = os.path.join(tmpdir, "queue.xlsx")
        data = [
            ["", "https://www.youtube.com/watch?v=dQw4w9WgXcQ", "Done", 119, "2026-07-16"],
            ["", "https://www.youtube.com/watch?v=abc123xyzAB", "In Progress", 29, "2026-07-16"],
            ["", "https://www.youtube.com/watch?v=def456uvwCD", "Not Started", None, ""],
        ]
        create_temp_xlsx(temp_excel, REQUIRED_COLUMNS, data)

        queue = VideoQueueManager(temp_excel)
        next_vid, warnings = queue.get_next_video()

        assert next_vid is not None
        assert next_vid["video_id"] == "abc123xyzAB"
        assert next_vid["status"] == "In Progress"
        assert next_vid["last_sec_completed"] == 29
        assert not warnings

        # The missing video ID should have been auto-filled in the sheet
        wb = openpyxl.load_workbook(temp_excel)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "dQw4w9WgXcQ"
        assert ws.cell(row=3, column=1).value == "abc123xyzAB"
        wb.close()


def test_get_next_video_duplicates():
    with TemporaryDirectory() as tmpdir:
        temp_excel = os.path.join(tmpdir, "queue.xlsx")
        data = [
            ["dQw4w9WgXcQ", "https://www.youtube.com/watch?v=dQw4w9WgXcQ", "Not Started", None, ""],
            ["dQw4w9WgXcQ", "https://www.youtube.com/watch?v=dQw4w9WgXcQ", "Not Started", None, ""],
            ["def456uvwCD", "https://www.youtube.com/watch?v=def456uvwCD", "Not Started", None, ""],
        ]
        create_temp_xlsx(temp_excel, REQUIRED_COLUMNS, data)

        queue = VideoQueueManager(temp_excel)
        next_vid, warnings = queue.get_next_video()

        assert next_vid is not None
        assert next_vid["video_id"] == "dQw4w9WgXcQ"
        assert len(warnings) == 1
        assert "Duplicate Video IDs detected" in warnings[0]

        # Mark done
        queue.mark_done("dQw4w9WgXcQ")

        # Get next should skip duplicate and return def456uvwCD
        next_vid2, warnings2 = queue.get_next_video()
        assert next_vid2 is not None
        assert next_vid2["video_id"] == "def456uvwCD"


def test_checkpoint_cross_check(monkeypatch):
    with TemporaryDirectory() as tmpdir:
        # Override project root directories / sessions dir for the cross-check test
        temp_sessions_dir = Path(tmpdir) / "data" / "sessions"
        temp_sessions_dir.mkdir(parents=True)
        
        # Write dummy checkpoint file
        checkpoint_file = temp_sessions_dir / "abc123xyzAB.json"
        checkpoint_data = {
            "session_info": {
                "video_id": "abc123xyzAB",
                "thresholds": {"T_body": 8.0, "T_head": 0.04, "T_hand": 0.03, "T_leg": 0.02}
            },
            "windows": [
                {
                    "window_id": 0,
                    "observations": [
                        {"time_sec": i, "hand": 0, "leg": 0, "head": 0, "body": 0} for i in range(10)
                    ]
                },
                {
                    "window_id": 1,
                    "observations": [
                        {"time_sec": i, "hand": 0, "leg": 0, "head": 0, "body": 0} for i in range(10, 20)
                    ]
                }
            ]
        }
        with open(checkpoint_file, "w") as f:
            json.dump(checkpoint_data, f)

        # Mock the path in video_queue to point to our temp sessions folder
        def mock_sessions_path(video_id):
            return temp_sessions_dir / f"{video_id}.json"

        # Create queue manager
        temp_excel = os.path.join(tmpdir, "queue.xlsx")
        create_temp_xlsx(temp_excel, REQUIRED_COLUMNS, [])
        queue = VideoQueueManager(temp_excel)

        # Monkeypatch Path in cross_check_checkpoint to use our mock paths
        original_cross_check = queue.cross_check_checkpoint
        def patched_cross_check(video_id, last_sec):
            # Temporarily replace Path constructor or just patch the local method logic.
            # Let's override the session_dir lookup inside Path(Path(__file__).parent.parent...)
            # Actually, we can patch the checkpoint path determination in cross_check_checkpoint by setting a local property or monkeypatch.
            # Let's inspect the method:
            # session_dir = Path(__file__).parent.parent / "data" / "sessions"
            # checkpoint_path = session_dir / f"{video_id}.json"
            # We can mock the checkpoint file check. Let's do it using monkeypatch on open/exists
            return original_cross_check(video_id, last_sec)
            
        # Instead of patching Path, let's create the folder structure relative to video_queue.py!
        # The session_dir inside video_queue is:
        # Path(__file__).parent.parent / "data" / "sessions"
        # Since Path(__file__).parent is e:\projects\internship_iit\auto_annotator,
        # session_dir is e:\projects\internship_iit\data\sessions.
        # This directory already exists in our workspace! We listed it earlier.
        # So we can write dummy json files to the real data/sessions directory during the test,
        # and delete them when the test completes. This is completely standard and guarantees compatibility.
        
        real_sessions_dir = Path(__file__).parent.parent / "auto_annotator" / "parent" / "data" / "sessions"
        # Wait, the relative path is:
        # e:\projects\internship_iit\auto_annotator\video_queue.py
        # Path(__file__).parent.parent is e:\projects\internship_iit
        # So the real sessions dir is e:\projects\internship_iit\data\sessions.
        # Let's verify: yes! e:\projects\internship_iit\data\sessions exists.
        # So we can just create a test checkpoint file in the real sessions dir:
        test_checkpoint = Path(__file__).parent.parent / "data" / "sessions" / "test_check_video.json"
        
        # 1. Test when Excel says Not Started (None) and checkpoint doesn't exist
        is_ok, msg, data = queue.cross_check_checkpoint("test_check_video", None)
        assert is_ok
        assert not msg

        # 2. Test when Excel says In Progress (19) but checkpoint doesn't exist
        is_ok, msg, data = queue.cross_check_checkpoint("test_check_video", 19)
        assert not is_ok
        assert "missing" in msg

        # 3. Create checkpoint with last second 19
        test_checkpoint.parent.mkdir(parents=True, exist_ok=True)
        checkpoint_content = {
            "session_info": {"video_id": "test_check_video"},
            "windows": [
                {
                    "window_id": 0,
                    "observations": [{"time_sec": i, "hand": 0, "leg": 0, "head": 0, "body": 0} for i in range(10)]
                },
                {
                    "window_id": 1,
                    "observations": [{"time_sec": i, "hand": 0, "leg": 0, "head": 0, "body": 0} for i in range(10, 20)]
                }
            ]
        }
        with open(test_checkpoint, "w", encoding="utf-8") as f:
            json.dump(checkpoint_content, f)

        try:
            # 4. Test agreement
            is_ok, msg, data = queue.cross_check_checkpoint("test_check_video", 19)
            assert is_ok
            assert not msg
            
            # 5. Test disagreement (Excel says 29, checkpoint says 19)
            is_ok, msg, data = queue.cross_check_checkpoint("test_check_video", 29)
            assert not is_ok
            assert "completed second" in msg
        finally:
            # Clean up
            if test_checkpoint.exists():
                test_checkpoint.unlink()


def test_mark_progress_and_done():
    with TemporaryDirectory() as tmpdir:
        temp_excel = os.path.join(tmpdir, "queue.xlsx")
        data = [
            ["abc123xyzAB", "https://www.youtube.com/watch?v=abc123xyzAB", "Not Started", None, ""],
        ]
        create_temp_xlsx(temp_excel, REQUIRED_COLUMNS, data)

        queue = VideoQueueManager(temp_excel)
        
        # Mark In Progress
        queue.mark_progress("abc123xyzAB", 19)
        
        # Verify
        wb = openpyxl.load_workbook(temp_excel)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "In Progress"
        assert ws.cell(row=2, column=4).value == 19
        assert ws.cell(row=2, column=5).value != ""
        wb.close()
        
        # Mark Done
        queue.mark_done("abc123xyzAB")
        
        wb = openpyxl.load_workbook(temp_excel)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "Done"
        wb.close()
